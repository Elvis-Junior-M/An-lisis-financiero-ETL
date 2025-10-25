from openpyxl import load_workbook                 # Para trabajar con Excel
from openpyxl.workbook.workbook import Workbook    # Tipo del libro
from openpyxl.worksheet.worksheet import Worksheet # Tipo de hoja
import sys
import pandas as pd
from tkinter import filedialog, Tk

# obtiene la ruta del archivo excel

def capturar_ruta(code:int) -> str:
    titulo: str = ""

    if code == 0:
        titulo = "Seleccionar archivo para extraer los datos"
    else:
         titulo = "Seleccionar archivo para cargar los datos"

    root = Tk()
    root.withdraw()
    ruta = filedialog.askopenfilename(
        title = titulo,
        filetypes = [("Archivos Excel","*.xlsx"), ("Todos los archivos","*.*")] #Solo trabaja con .xlsx, .xlsm, .xltx, .xltm
    )
    if not ruta:  # si está vacío
        print("No se seleccionó ningún archivo. El programa finalizará.")
        sys.exit(0)  # termina el programa
    return ruta


# cargamos los datos extraidos del reporte en el dataframe

def cargar_datos_extra(recibir_ruta:str) -> pd.DataFrame:

    # Declaramos el dataframe y variables
    nombre_hoja : str = "Original"
    ultima_fila : int = 0
    fila : int
    df_contenido: pd.DataFrame = None

    # Objeto que representa esa libro en memoria
    libro_extraer: Workbook = load_workbook(recibir_ruta,data_only=True) # Objeto Workbook
    
    # Verificar si la hoja existe
    if nombre_hoja not in libro_extraer:
        print(f"La hoja '{nombre_hoja}' no existe")
        sys.exit(1)
    
    # Objeto que representa esa hoja en memoria
    hoja_extraer: Worksheet = libro_extraer[nombre_hoja]

    # Encontrar la última fila con datos en la columna B
    for fila in range(hoja_extraer.max_row, 0, -1):                     # desde la última fila hacia arriba
        valor_celda:object = hoja_extraer.cell(row=fila,column=2).value # valor de la celda
        if valor_celda is not None:
            ultima_fila = fila
            break
    
    # Cargar solo el rango deseado A6:L84
    df_contenido = pd.read_excel(
        recibir_ruta,
        sheet_name = nombre_hoja,
        usecols = "A:L",
        skiprows = 6-1,                 # Saltar hasta la fila 6  (-1) por que cuenta de 0
        nrows = ultima_fila - (6-1),    # Total de filas desde la 6, cantidad de filas exactas
        header=None,                    # no tomar fila como encabezado
        engine ="openpyxl"
    ) 
    return df_contenido

# cargamos los datos limpios de la hoja Auxiliar Balance al dataframe

def cargar_auxiliar_balance(recibir_ruta:str) -> pd.DataFrame:

    # Declaramos el dataframe y variables
    nombre_hoja : str = "Auxiliar Balance"
    ultima_fila : int = 0
    fila : int
    df_contenido_aux: pd.DataFrame = None

    # Cargar el libro Excel
    libro_extraer: Workbook = load_workbook(recibir_ruta,data_only=True) # Objeto Workbook
    
    # Verificar si la hoja existe
    if nombre_hoja not in libro_extraer:
        print(f"La hoja '{nombre_hoja}' no existe")
        sys.exit(1)
    
    # Seleccionar la hoja
    hoja_extraer: Worksheet = libro_extraer[nombre_hoja]

    # Encontrar la última fila con datos en la columna B
    for fila in range(hoja_extraer.max_row, 0, -1):                          # desde la última fila hacia arriba
        valor_celda:object = hoja_extraer.cell(row = fila, column = 2).value # valor de la celda
        if valor_celda is not None:
            ultima_fila = fila
            break
    
    # Cargar solo el rango deseado A1:E78
    df_contenido_aux = pd.read_excel(
        recibir_ruta,
        sheet_name = nombre_hoja,
        usecols = "A:E",
        skiprows = 0,                   # Saltar hasta la fila 6  (-1) por que cuenta de 0
        nrows = ultima_fila,            # Total de filas desde la 6, cantidad de filas exactas
        header=None,                    # no tomar fila como encabezado
        engine ="openpyxl"
    ) 
    return df_contenido_aux

# cargamos la tabla Clasificacion al dataframe =A4:B9

def cargar_clasi(recibir_ruta:str) -> pd.DataFrame:

    # Declaramos el dataframe y variables
    nombre_hoja : str = "Tablas"
    ultima_fila : int = 0
    fila : int
    df_contenido_clasi: pd.DataFrame = None

    # Cargar el libro Excel
    libro_clasi: Workbook = load_workbook(recibir_ruta,data_only=True) # Objeto Workbook
    
    # Verificar si la hoja existe
    if nombre_hoja not in libro_clasi:
        print(f"La hoja '{nombre_hoja}' no existe")
        sys.exit(1)
    
    # Seleccionar la hoja
    hoja_tabla: Worksheet = libro_clasi[nombre_hoja]

    # Encontrar la última fila con datos en la columna B
    for fila in range(hoja_tabla.max_row, 0, -1):                          # desde la última fila hacia arriba
        valor_celda:object = hoja_tabla.cell(row = fila, column = 1).value 
        if valor_celda is not None:
            ultima_fila = fila
            break
    
    # Cargar solo el rango deseado
    df_contenido_clasi = pd.read_excel(
        recibir_ruta,
        sheet_name = nombre_hoja,
        usecols = "A:B",
        skiprows = 3,                   # saltar filas 0 a 3 -> comenzar desde la 4
        nrows = ultima_fila - 4 + 1,    # cantidad de filas = ultima_fila - fila_inicio + 1
        header=None,                    # no tomar fila como encabezado
        engine ="openpyxl"
    ) 
    print(df_contenido_clasi)
    return df_contenido_clasi

# cargamos la tabla Tipo al dataframe =D4:E13

def cargar_tipo(recibir_ruta:str) -> pd.DataFrame:

    # Declaramos el dataframe y variables
    nombre_hoja : str = "Tablas"
    ultima_fila : int = 0
    fila : int
    df_contenido_tipo: pd.DataFrame = None

    # Cargar el libro Excel
    libro_tipo: Workbook = load_workbook(recibir_ruta,data_only=True) # Objeto Workbook
    
    # Verificar si la hoja existe
    if nombre_hoja not in libro_tipo:
        print(f"La hoja '{nombre_hoja}' no existe")
        sys.exit(1)
    
    # Seleccionar la hoja
    hoja_tabla: Worksheet = libro_tipo[nombre_hoja]

    # Encontrar la última fila con datos en la columna B
    for fila in range(hoja_tabla.max_row, 0, -1):                          # desde la última fila hacia arriba
        valor_celda:object = hoja_tabla.cell(row = fila, column = 4).value 
        if valor_celda is not None:
            ultima_fila = fila
            break
    
    # Cargar solo el rango deseado A4:B9
    df_contenido_tipo = pd.read_excel(
        recibir_ruta,
        sheet_name = nombre_hoja,
        usecols = "D:E",
        skiprows = 3,                   # saltar filas 0 a 3 -> comenzar desde la 4
        nrows = ultima_fila - 4 + 1,    # cantidad de filas = ultima_fila - fila_inicio + 1
        header=None,                    # no tomar fila como encabezado
        engine ="openpyxl"
    )
    print(df_contenido_tipo) 
    return df_contenido_tipo

# cargamos la tabla Detalle al dataframe =G4:H59

def cargar_detalle(recibir_ruta:str) -> pd.DataFrame:

    # Declaramos el dataframe y variables
    nombre_hoja : str = "Tablas"
    ultima_fila : int = 0
    fila : int
    df_contenido_deta: pd.DataFrame = None

    # Cargar el libro Excel
    libro_deta: Workbook = load_workbook(recibir_ruta,data_only=True) # Objeto Workbook
    
    # Verificar si la hoja existe
    if nombre_hoja not in libro_deta:
        print(f"La hoja '{nombre_hoja}' no existe")
        sys.exit(1)
    
    # Seleccionar la hoja
    hoja_tabla: Worksheet = libro_deta[nombre_hoja]

    # Encontrar la última fila con datos en la columna B
    for fila in range(hoja_tabla.max_row, 0, -1):                          # desde la última fila hacia arriba
        valor_celda:object = hoja_tabla.cell(row = fila, column = 7).value 
        if valor_celda is not None:
            ultima_fila = fila
            break
    
    # Cargar solo el rango deseado A4:B9
    df_contenido_deta = pd.read_excel(
        recibir_ruta,
        sheet_name = nombre_hoja,
        usecols = "G:H",
        skiprows = 3,                   # saltar filas 0 a 3 -> comenzar desde la 4
        nrows = ultima_fila - 4 + 1,    # cantidad de filas = ultima_fila - fila_inicio + 1
        header=None,                    # no tomar fila como encabezado
        engine ="openpyxl"
    ) 
    print(df_contenido_deta)
    return df_contenido_deta