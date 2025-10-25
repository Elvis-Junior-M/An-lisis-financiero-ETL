from openpyxl import load_workbook                 # Para trabajar con Excel
from openpyxl.workbook.workbook import Workbook    # Tipo del libro
from openpyxl.worksheet.worksheet import Worksheet # Tipo de hoja
from openpyxl.cell.cell import Cell
from datetime import datetime
from transformar import buscar_id
import pandas as pd
import sys

def registrar(libro_car:Workbook, fecha:str, datos_lim:list[str], valor:float, df_cla: pd.DataFrame, df_tip: pd.DataFrame, df_det: pd.DataFrame) -> Workbook:
    # Declaramos variables
    hoja : str = "BASE DATOS"
    ultima_fila : int = 0
    fila : int
    fecha_conv : datetime
    i: int = 0
    id_clasi: int = 0
    id_tipo: int = 0
    id_detalle: int = 0
    dataframes = {}

    # Verificar si la hoja existe
    if hoja not in libro_car:
        print(f"La hoja '{hoja}' no existe")
        sys.exit(1)

    # Objeto que representa esa hoja en memoria
    hoja_DB: Worksheet = libro_car[hoja]

    # Llenamos el diccionario con los dataframen
    dataframes = {
        "cla": df_cla,
        "tip": df_tip,
        "det": df_det
    }

    # Encontrar la última fila con datos en la columna B
    for fila in range(hoja_DB.max_row, 0, -1):                     # desde la última fila hacia arriba
        valor_celda:object = hoja_DB.cell(row=fila,column=2).value # valor de la celda
        if valor_celda is not None:
            ultima_fila = fila
            break

    # Convertir texto a fecha
    fecha_conv = datetime.strptime(fecha, "%d/%m/%Y")
    # Definir objeto celda
    celda: Cell = hoja_DB.cell(row = fila + 1, column = 1, value= fecha_conv)  # FECHA
    celda.number_format = "DD/MM/YYY"

    # lista [nombre_lim, clasificacion, tipo]
    hoja_DB.cell(row = fila + 1, column = 2, value= str(datos_lim[1]))  # CLASIFICACION
    hoja_DB.cell(row = fila + 1, column = 3, value= str(datos_lim[2]))  # TIPO
    hoja_DB.cell(row = fila + 1, column = 4, value= str(datos_lim[0]))  # DETALLE
    hoja_DB.cell(row = fila + 1, column = 5, value= valor)  # VALOR

    # Buscamos los ids
    for df_nombre, df in dataframes.items():
        match df_nombre:
            case "cla":
                id_clasi = buscar_id(df_cla, str(datos_lim[1]))
                hoja_DB.cell(row = fila + 1, column = 6, value = id_clasi)
            case "tip":
                id_tipo = buscar_id(df_tip, str(datos_lim[2]))
                hoja_DB.cell(row = fila + 1, column = 7, value = id_tipo)
            case "det":
                id_detalle = buscar_id(df_det, str(datos_lim[0]))
                hoja_DB.cell(row = fila + 1, column = 8, value = id_detalle)

    return libro_car

# Guardamos el libro con todos los cambios

def guardar_libro(ruta: str ,libro_carg: Workbook, registro: int) -> bool:
    try:
        # Guardamos los cambios
        libro_carg.save(ruta)
        print(f"Se han registrado {registro} registros en la hoja Base de Datos.")
        return True
    except PermissionError:
        print("❌ No se puede guardar: el archivo está abierto en Excel.")
        return False
    except FileNotFoundError:
        print("❌ No se encontró el archivo Excel en la ruta indicada.")
        return False
    except Exception as e:
        print(f"⚠️ Error inesperado al guardar: {e}")
        return False
    
# Extiende la tabla especificada hasta la última fila con datos
# para que se aplique el formato de la tabla
    
def formato_tabla(libro_carga: Workbook) -> Workbook:
    hoja : str = "BASE DATOS"
    tabla : str = "DB_BALANCE"
    ultima_fila: int = 0

    # Verificar si la hoja existe
    if hoja not in libro_carga:
        print(f"La hoja '{hoja}' no existe")
        sys.exit(1)
    
    # Objeto que representa esa hoja en memoria
    hoja_DB: Worksheet = libro_carga[hoja]

    # Verificar que la tabla exista
    if tabla  not in hoja_DB.tables:
         print(f"⚠️ No se encontró la tabla '{tabla}' en la hoja.")
         sys.exit(1)

    # Buscamos la tabla
    tabla_DB = hoja_DB.tables[tabla]

    # Buscar la última fila con datos en la columna B
    for fila in range(hoja_DB.max_row, 0, -1):                # Recorre de abajo hacia arriba
        valor_celda = hoja_DB.cell(row=fila, column=2).value  # columna B = 2
        if valor_celda is not None:
            ultima_fila = fila
            break
    
    tabla_DB.ref = f"A3:H{ultima_fila}"

    return libro_carga